from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from ghed_mcp.client import (
    document_download_url,
    find_latest_all_data_document,
    normalize_source_document,
)
from ghed_mcp import server


@pytest.fixture
def sample_workbook(tmp_path: Path) -> Path:
    path = tmp_path / "ghed.xlsx"
    wb = Workbook()

    ws = wb.active
    ws.title = "Data"
    ws.append([
        "location", "code", "region", "income", "year",
        "che_gdp", "oops_che", "che", "gghed", "fs1", "fs3", "hc", "hc1", "hc2",
    ])
    ws.append([
        "Colombia", "COL", "AMR", "Upper-middle", 2022,
        8.1, 14.0, 100.0, 70.0, 60.0, 10.0, 100.0, 55.0, 45.0,
    ])
    ws.append([
        "Colombia", "COL", "AMR", "Upper-middle", 2023,
        8.3, 13.8, 110.0, 80.0, 65.0, 15.0, 110.0, 60.0, 50.0,
    ])
    ws.append([
        "Peru", "PER", "AMR", "Upper-middle", 2022,
        5.5, 28.0, 90.0, 50.0, 45.0, 5.0, 90.0, 40.0, 50.0,
    ])
    ws.append([
        "United States of America", "USA", "AMR", "High", 2023,
        16.5, 10.0, 1000.0, 500.0, 400.0, 100.0, 1000.0, 650.0, 350.0,
    ])

    ws = wb.create_sheet("Codebook")
    ws.append([
        "variable code",
        "variable name",
        "long code (GHED data explorer)",
        "category 1",
        "category 2",
        "unit",
        "currency",
        "Method of measurement (INDICATORS category1)",
    ])
    ws.append(["location", "Countries and territories name", "-", "-", "-", "-", "-", "-"])
    ws.append([
        "che_gdp",
        "Current Health Expenditure (CHE) as % Gross Domestic Product (GDP)",
        "CHE%GDP_SHA2011",
        "INDICATORS",
        "AGGREGATES",
        "Percentage",
        "-",
        "CHE / GDP",
    ])
    ws.append([
        "oops_che",
        "Out-of-pocket spending as % of current health expenditure",
        "OOPS%CHE_SHA2011",
        "INDICATORS",
        "FINANCING",
        "Percentage",
        "-",
        "OOPS / CHE",
    ])
    ws.append([
        "hc11",
        "Inpatient curative care",
        "sha11.HC.1.1",
        "HEALTH EXPENDITURE DATA",
        "HEALTH CARE FUNCTIONS",
        "National Currency Unit (NCU) millions",
        "-",
        "-",
    ])
    ws.append([
        "che",
        "Current Health Expenditure (CHE), in million current NCU",
        "CHE",
        "INDICATORS",
        "AGGREGATES",
        "Millions",
        "NCU",
        "HF.1 + HF.2 + HF.3 + HF.4 + HF.nec",
    ])
    ws.append([
        "gghed",
        "Domestic General Government Health Expenditure (GGHE-D), in million current NCU",
        "GGHE-D",
        "INDICATORS",
        "AGGREGATES",
        "Millions",
        "NCU",
        "FS.1 + FS.3",
    ])
    ws.append([
        "fs1",
        "Transfers from government domestic revenue",
        "sha11.FS.1",
        "HEALTH EXPENDITURE DATA",
        "REVENUES",
        "Millions",
        "NCU",
        "-",
    ])
    ws.append([
        "fs3",
        "Social insurance contributions",
        "sha11.FS.3",
        "HEALTH EXPENDITURE DATA",
        "REVENUES",
        "Millions",
        "NCU",
        "-",
    ])
    ws.append([
        "hc",
        "Current health expenditure by Health Care Functions",
        "sha11.HC",
        "HEALTH EXPENDITURE DATA",
        "HEALTH CARE FUNCTIONS",
        "Millions",
        "NCU",
        "-",
    ])
    ws.append([
        "hc1",
        "Curative care",
        "sha11.HC.1",
        "HEALTH EXPENDITURE DATA",
        "HEALTH CARE FUNCTIONS",
        "Millions",
        "NCU",
        "-",
    ])
    ws.append([
        "hc2",
        "Rehabilitative care",
        "sha11.HC.2",
        "HEALTH EXPENDITURE DATA",
        "HEALTH CARE FUNCTIONS",
        "Millions",
        "NCU",
        "-",
    ])

    ws = wb.create_sheet("Metadata")
    ws.append([
        "location",
        "code",
        "region",
        "income",
        "variable code",
        "long code (GHED data explorer)",
        "variable name",
        "Sources",
        "Comments",
        "Data type",
        "Methods of estimation",
        "Countries and territories footnote",
    ])
    ws.append([
        "Colombia",
        "COL",
        "AMR",
        "Upper-middle",
        "che_gdp",
        "CHE%GDP_SHA2011",
        "Current Health Expenditure (CHE) as % Gross Domestic Product (GDP)",
        "Official health accounts",
        None,
        "Documented",
        None,
        "Calendar year",
    ])

    ws = wb.create_sheet("Version")
    ws.append(["WHO Global Health Expenditure Database (GHED)"])
    ws.append(["Last updated: test"])

    wb.save(path)
    return path


@pytest.fixture(autouse=True)
def fake_store(monkeypatch, sample_workbook):
    async def _fake_get_store(refresh: bool = False):
        return server.GHEDStore(sample_workbook)

    monkeypatch.setattr(server, "get_store", _fake_get_store)


async def test_search_indicators():
    result = await server.search_indicators("out-of-pocket")

    assert result["count"] == 1
    assert result["items"][0]["indicator_code"] == "oops_che"


async def test_list_indicators_is_headline_only():
    result = await server.list_indicators()

    assert result["total"] == 4
    assert {row["indicator_code"] for row in result["items"]} == {
        "che_gdp",
        "che",
        "gghed",
        "oops_che",
    }


async def test_list_variables_includes_detailed_series():
    result = await server.list_variables(category_1="HEALTH EXPENDITURE DATA")

    assert result["total"] == 6
    assert "hc11" in {row["indicator_code"] for row in result["items"]}


async def test_methodology_guide_has_variable_classes():
    result = await server.methodology_guide()

    assert "INDICATORS" in result["summary"]["variable_classes"]
    assert result["current_counts"]["category_1"][0]["count"] >= 1


async def test_suggest_variables_for_research_question():
    result = await server.suggest_variables_for_research_question(
        "Did out-of-pocket spending decline after government spending rose?"
    )

    use_cases = [item["use_case"] for item in result["suggestions"]]
    assert "financial_protection_oop" in use_cases


async def test_suggest_variables_includes_alias_matches():
    result = await server.suggest_variables_for_research_question(
        "Which countries have the highest fiscal priority?"
    )

    assert {
        "indicator_code": "gghed_gge",
        "matched_aliases": ["fiscal priority"],
    } in result["matched_indicator_aliases"]


async def test_data_availability_for_panel_inputs():
    result = await server.data_availability(
        ["che_gdp", "oops_che"],
        countries=["Colombia", "Peru"],
        year_start=2022,
        year_end=2023,
    )

    by_code = {row["indicator_code"]: row for row in result["items"]}
    assert by_code["che_gdp"]["observation_count"] == 3
    assert by_code["che_gdp"]["country_count"] == 2
    assert by_code["oops_che"]["latest_year"] == 2023


async def test_country_groups_and_filters():
    groups = await server.list_country_groups()
    assert {"region": "AMR", "country_count": 3} in groups["regions"]
    assert {"income": "Upper-middle", "country_count": 2} in groups["income_groups"]

    upper_middle = await server.list_countries(income="Upper-middle")
    assert upper_middle["count"] == 2
    assert {row["country_code"] for row in upper_middle["items"]} == {"COL", "PER"}


async def test_data_availability_accepts_region_and_income_filters():
    result = await server.data_availability(
        ["che_gdp"],
        region="AMR",
        income="High",
    )

    assert result["region"] == "AMR"
    assert result["income"] == "High"
    assert result["items"][0]["country_count"] == 1
    assert result["items"][0]["observation_count"] == 1


async def test_build_research_panel_csv():
    result = await server.build_research_panel(
        ["che_gdp", "oops_che"],
        countries=["Colombia"],
        year_start=2022,
        year_end=2023,
        format="csv",
    )

    assert result["count"] == 4
    assert result["csv"].startswith("indicator_code,indicator_name")
    assert "che_gdp" in result["csv"]
    assert "oops_che" in result["csv"]


async def test_build_research_panel_accepts_group_filters():
    result = await server.build_research_panel(
        ["che_gdp"],
        region="AMR",
        income="High",
    )

    assert result["count"] == 1
    assert result["rows"][0]["country_code"] == "USA"


async def test_compare_countries_latest_csv():
    result = await server.compare_countries(
        "che_gdp",
        ["Colombia", "PER"],
        latest_only=True,
        format="csv",
    )

    assert result["countries_resolved"] == [
        {"input": "Colombia", "code": "COL"},
        {"input": "PER", "code": "PER"},
    ]
    assert "csv" in result
    assert "COL,Colombia,AMR,Upper-middle,2023,8.3" in result["csv"]
    assert "PER,Peru,AMR,Upper-middle,2022,5.5" in result["csv"]
    assert result["warnings"][0]["type"] == "mixed_latest_years"


async def test_compare_country_group():
    result = await server.compare_country_group(
        "che_gdp",
        income="Upper-middle",
        latest_only=True,
    )

    assert result["country_count"] == 2
    assert {row["country_code"] for row in result["countries"]} == {"COL", "PER"}
    assert {row["country_code"] for row in result["rows"]} == {"COL", "PER"}
    assert result["warnings"][0]["type"] == "mixed_latest_years"


async def test_summarize_country_group():
    result = await server.summarize_country_group(
        "che_gdp",
        income="Upper-middle",
        latest_only=True,
    )

    assert result["stats"]["n"] == 2
    assert result["stats"]["median"] == 6.9
    assert result["coverage_ratio"] == 1.0
    assert result["top"][0]["country_code"] == "COL"
    assert result["warnings"][0]["type"] == "mixed_latest_years"


async def test_find_country_code_accepts_country_and_alias():
    via_canonical = await server.find_country_code(country="United States")
    via_deprecated_alias = await server.find_country_code(country_name="United States")
    via_common_alias = await server.find_country_code(country="US")

    assert via_canonical["matches"] == via_deprecated_alias["matches"]
    assert via_common_alias["matches"][0]["country_code"] == "USA"

    with pytest.raises(ValueError, match="not both"):
        await server.find_country_code(country="Colombia", country_name="Colombia")

    with pytest.raises(ValueError, match="'country' parameter"):
        await server.find_country_code()


async def test_compare_countries_resolves_country_aliases():
    result = await server.compare_countries(
        "che_gdp",
        ["US"],
        latest_only=True,
    )

    assert result["countries_resolved"] == [{"input": "US", "code": "USA"}]
    assert result["rows"][0]["country_name"] == "United States of America"


async def test_tool_arguments_reject_unknown_names():
    from mcp.server.fastmcp.exceptions import ToolError

    with pytest.raises(ToolError, match="Extra inputs are not permitted"):
        await server.mcp._tool_manager.call_tool(
            "get_indicator_data",
            {
                "indicator_code": "che_gdp",
                "countries": ["Peru"],
                "year_from": 2022,
            },
        )


async def test_get_country_metadata():
    result = await server.get_country_metadata(country="Colombia", indicator_code="che_gdp")

    assert result["count"] == 1
    assert result["rows"][0]["sources"] == "Official health accounts"


async def test_country_profile_uses_default_indicators():
    result = await server.country_profile("COL", indicator_codes=["che_gdp", "oops_che"])

    assert result["country_code"] == "COL"
    assert [row["indicator_code"] for row in result["indicators"]] == [
        "che_gdp",
        "oops_che",
    ]
    assert result["indicators"][0]["year"] == 2023


async def test_cache_status_reports_sqlite_counts():
    result = await server.cache_status()

    assert result["cache"]["sqlite_current"] is True
    assert result["cache"]["counts"]["countries"] == 3
    assert result["cache"]["counts"]["indicators"] == 10
    assert result["cache"]["counts"]["observations"] == 36


async def test_additive_hierarchy_formula_and_sha_children():
    formula = await server.additive_hierarchy("gghed")
    assert formula["relationships"][0]["children"] == ["fs1", "fs3"]

    sha = await server.additive_hierarchy("hc")
    rel = {
        item["relationship_id"]: item for item in sha["relationships"]
    }["sha_direct_children"]
    assert rel["children"] == ["hc1", "hc2"]


async def test_explain_indicator_relationship():
    parent = await server.explain_indicator_relationship("gghed")
    child = await server.explain_indicator_relationship("fs1")
    share = await server.explain_indicator_relationship("che_gdp")

    assert parent["role"] == "additive_parent"
    assert parent["additive_children"][0]["children"] == ["fs1", "fs3"]
    assert child["role"] == "component"
    assert child["known_parents"][0]["parent_code"] == "gghed"
    assert share["role"] == "derived_ratio_or_share"


async def test_build_additive_breakdown_balances():
    result = await server.build_additive_breakdown(
        "gghed",
        country="Colombia",
        year=2023,
    )

    assert result["parent"]["value"] == 80.0
    assert result["child_sum"] == 80.0
    assert result["balanced"] is True
    assert [row["indicator_code"] for row in result["children"]] == ["fs1", "fs3"]


async def test_build_research_panel_warns_when_top_limit_reached():
    result = await server.build_research_panel(["che_gdp"], top=2)

    assert result["possibly_truncated"] is True
    assert result["warnings"][0]["type"] == "top_limit_reached"


async def test_trend_and_rank_tools():
    trend = await server.indicator_trend(
        "che_gdp",
        countries=["Colombia", "Peru"],
        year_start=2022,
        year_end=2023,
    )
    by_code = {row["country_code"]: row for row in trend["rows"]}
    assert by_code["COL"]["absolute_change"] == pytest.approx(0.2)
    assert by_code["PER"]["year_count"] == 1

    ranked = await server.rank_country_changes(
        "che_gdp",
        countries=["Colombia", "Peru"],
        year_start=2022,
        year_end=2023,
    )
    assert ranked["rows"][0]["country_code"] == "COL"

    compared = await server.compare_trends(
        ["che_gdp", "oops_che"],
        countries=["Colombia"],
    )
    assert [item["indicator_code"] for item in compared["items"]] == [
        "che_gdp",
        "oops_che",
    ]


async def test_assess_data_quality():
    result = await server.assess_data_quality("che_gdp", country="Colombia")

    assert result["availability"]["observation_count"] == 2
    assert result["metadata_summary"]["data_types"] == {"Documented": 1}
    assert result["metadata_summary"]["rows_with_country_footnotes"] == 1
    assert result["warnings"][0]["type"] == "country_notes_present"


async def test_build_research_package():
    result = await server.build_research_package(
        ["che_gdp", "oops_che"],
        countries=["Colombia"],
        year_start=2022,
        year_end=2023,
    )

    assert result["count"] == 4
    assert result["data_csv"].startswith("indicator_code,indicator_name")
    assert "che_gdp" in result["codebook_csv"]
    assert "GHED Research Extract" in result["readme"]


async def test_topic_and_research_use_case_resources():
    topic = await server.topic_resource("out_of_pocket")
    use_case = await server.research_use_case_resource("financial_protection_oop")
    missing = await server.topic_resource("missing")

    assert "# out_of_pocket" in topic
    assert "`oops_che`" in topic
    assert "# financial_protection_oop" in use_case
    assert "Unknown GHED topic" in missing


def test_find_latest_all_data_document_uses_documentation_tree():
    tree = {
        "IsFolder": True,
        "Children": [
            {
                "IsFolder": True,
                "Name": "Download GHED all data",
                "Children": [
                    {
                        "IsFolder": False,
                        "Identifier": 1,
                        "Name": "GHED all data (December 2025)",
                        "FileType": ".xlsx",
                        "DateModified": "/Date(1765790273000)/",
                    },
                    {
                        "IsFolder": False,
                        "Identifier": 64396441,
                        "Name": "GHED all data (March 2026)",
                        "FileType": ".xlsx",
                        "DateModified": "/Date(1774900345000)/",
                    },
                ],
            }
        ],
    }

    doc = find_latest_all_data_document(tree)

    assert doc["Identifier"] == 64396441
    assert document_download_url(doc["Identifier"]).endswith(
        "/DocumentationCentre/GetFile/64396441/en"
    )


async def test_region_alias_resolution():
    via_alias = await server.compare_country_group(
        "che_gdp",
        region="Americas",
        latest_only=True,
    )
    via_canonical = await server.compare_country_group(
        "che_gdp",
        region="AMR",
        latest_only=True,
    )

    assert {row["country_code"] for row in via_alias["rows"]} == {
        row["country_code"] for row in via_canonical["rows"]
    }


async def test_income_alias_resolution():
    via_canonical = await server.list_countries(income="Upper-middle")
    via_natural = await server.list_countries(income="upper middle income")
    via_short = await server.list_countries(income="UMIC")

    expected = {row["country_code"] for row in via_canonical["items"]}
    assert {row["country_code"] for row in via_natural["items"]} == expected
    assert {row["country_code"] for row in via_short["items"]} == expected


async def test_unknown_region_lists_available_values():
    with pytest.raises(ValueError, match="Available regions"):
        await server.list_countries(region="Atlantis")


async def test_unknown_income_lists_available_values():
    with pytest.raises(ValueError, match="Available income groups"):
        await server.list_countries(income="middle-class")


async def test_lmic_is_not_aliased_to_lower_middle():
    # "LMIC" in global health usage means low + middle income countries
    # collectively, not just lower-middle. The alias map intentionally omits
    # it so researchers can't silently exclude Low and Upper-middle groups.
    with pytest.raises(ValueError, match="Available income groups"):
        await server.list_countries(income="LMIC")


async def test_indicator_trend_carries_period_years():
    trend = await server.indicator_trend(
        "che_gdp",
        countries=["Colombia", "Peru"],
        year_start=2022,
        year_end=2023,
    )

    by_code = {row["country_code"]: row for row in trend["rows"]}
    assert by_code["COL"]["period_years"] == 1
    assert by_code["PER"]["period_years"] == 0


async def test_rank_country_changes_filters_by_min_year_count():
    ranked = await server.rank_country_changes(
        "che_gdp",
        countries=["Colombia", "Peru"],
        year_start=2022,
        year_end=2023,
        min_year_count=2,
    )

    assert {row["country_code"] for row in ranked["rows"]} == {"COL"}


def test_normalize_source_document():
    doc = normalize_source_document({
        "Identifier": 64396441,
        "Name": "GHED all data (March 2026)",
        "Description": "GHED all data (March 2026)",
        "FileType": ".xlsx",
        "FileName": "GHED all data (March 2026).xlsx",
        "FileSize": 38922209,
        "DateModified": "/Date(1774900345000)/",
    })

    assert doc["document_id"] == 64396441
    assert doc["date_modified"] == "2026-03-30T19:52:25Z"
    assert doc["download_url"].endswith("/DocumentationCentre/GetFile/64396441/en")
