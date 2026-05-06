"""Read the GHED workbook into a derived SQLite cache and query it."""
from __future__ import annotations

import csv
import io
import json
import logging
import re
import sqlite3
import statistics
import time
import unicodedata
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from .methodology import ADDITIVE_RELATIONSHIPS

logger = logging.getLogger("ghed_mcp.store")

CORE_COLUMNS = {"location", "code", "region", "income", "year"}
DEFAULT_PROFILE_INDICATORS = [
    "che_gdp",
    "che_pc_usd",
    "gghed_che",
    "oops_che",
    "ext_che",
    "gghed_gdp",
    "gghed_gge",
]
SCHEMA_VERSION = 1
COUNTRY_ALIASES: dict[str, str] = {
    "US": "USA",
    "U.S.": "USA",
    "U.S.A.": "USA",
    "USA": "USA",
    "UK": "GBR",
    "U.K.": "GBR",
    "BRITAIN": "GBR",
    "GREAT BRITAIN": "GBR",
    "RUSSIA": "RUS",
    "SOUTH KOREA": "KOR",
    "KOREA": "KOR",
    "NORTH KOREA": "PRK",
    "IRAN": "IRN",
    "VIETNAM": "VNM",
    "VIET NAM": "VNM",
    "VENEZUELA": "VEN",
    "BOLIVIA": "BOL",
    "TANZANIA": "TZA",
    "DRC": "COD",
    "DR CONGO": "COD",
    "CZECHIA": "CZE",
    "CZECH REPUBLIC": "CZE",
    "TURKEY": "TUR",
    "TÜRKIYE": "TUR",
    "TURKIYE": "TUR",
}

REGION_ALIASES: dict[str, str] = {
    "AFR": "AFR",
    "AFRO": "AFR",
    "AFRICA": "AFR",
    "AFRICAN": "AFR",
    "AFRICAN REGION": "AFR",
    "WHO AFRICAN REGION": "AFR",
    "AMR": "AMR",
    "AMRO": "AMR",
    "AMERICAS": "AMR",
    "AMERICA": "AMR",
    "REGION OF THE AMERICAS": "AMR",
    "PAHO": "AMR",
    "EMR": "EMR",
    "EMRO": "EMR",
    "EASTERN MEDITERRANEAN": "EMR",
    "EASTERN MEDITERRANEAN REGION": "EMR",
    "EUR": "EUR",
    "EURO": "EUR",
    "EUROPE": "EUR",
    "EUROPEAN": "EUR",
    "EUROPEAN REGION": "EUR",
    "SEAR": "SEAR",
    "SEARO": "SEAR",
    "SOUTH EAST ASIA": "SEAR",
    "SOUTHEAST ASIA": "SEAR",
    "SOUTH EAST ASIA REGION": "SEAR",
    "WPR": "WPR",
    "WPRO": "WPR",
    "WESTERN PACIFIC": "WPR",
    "WESTERN PACIFIC REGION": "WPR",
}

INCOME_ALIASES: dict[str, str] = {
    "LOW": "Low",
    "LOW INCOME": "Low",
    "LOW INCOME COUNTRIES": "Low",
    "LIC": "Low",
    "LOWER MIDDLE": "Lower-middle",
    "LOWER MIDDLE INCOME": "Lower-middle",
    "LOWER MIDDLE INCOME COUNTRIES": "Lower-middle",
    "UPPER MIDDLE": "Upper-middle",
    "UPPER MIDDLE INCOME": "Upper-middle",
    "UPPER MIDDLE INCOME COUNTRIES": "Upper-middle",
    "UMIC": "Upper-middle",
    "HIGH": "High",
    "HIGH INCOME": "High",
    "HIGH INCOME COUNTRIES": "High",
    "HIC": "High",
}

# Collective aliases that expand to multiple GHED income labels. In global
# health literature "LMICs" almost always refers to low + middle income
# countries collectively (Low + Lower-middle + Upper-middle), not the World
# Bank's narrower "lower-middle-income" definition. We resolve the academic
# collective interpretation; users who want only Lower-middle can pass
# "Lower-middle" or "lower middle income".
INCOME_GROUP_ALIASES: dict[str, list[str]] = {
    "LMIC": ["Low", "Lower-middle", "Upper-middle"],
    "LMICS": ["Low", "Lower-middle", "Upper-middle"],
    "LMC": ["Low", "Lower-middle", "Upper-middle"],
    "LOW AND MIDDLE INCOME": ["Low", "Lower-middle", "Upper-middle"],
    "LOW AND MIDDLE INCOME COUNTRIES": ["Low", "Lower-middle", "Upper-middle"],
    "NON HIC": ["Low", "Lower-middle", "Upper-middle"],
    "NON HIGH INCOME": ["Low", "Lower-middle", "Upper-middle"],
    "NON HIGH INCOME COUNTRIES": ["Low", "Lower-middle", "Upper-middle"],
    "MIC": ["Lower-middle", "Upper-middle"],
    "MICS": ["Lower-middle", "Upper-middle"],
    "MIDDLE INCOME": ["Lower-middle", "Upper-middle"],
    "MIDDLE INCOME COUNTRIES": ["Lower-middle", "Upper-middle"],
}


def _norm_group_key(value: str) -> str:
    """Normalize a region/income value into a lookup key.

    Strips accents, uppercases, and collapses hyphens/underscores/whitespace
    so that 'Upper-middle', 'upper middle income', and 'UMIC' all collide.
    """
    normalized = unicodedata.normalize("NFKD", value.strip())
    ascii_value = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    return re.sub(r"[\s\-_]+", " ", ascii_value.upper()).strip()


@dataclass(frozen=True)
class Indicator:
    indicator_code: str
    indicator_name: str | None
    long_code: str | None
    category_1: str | None
    category_2: str | None
    unit: str | None
    currency: str | None
    measurement_method: str | None

    def to_dict(self) -> dict[str, Any]:
        return {
            "indicator_code": self.indicator_code,
            "indicator_name": self.indicator_name,
            "long_code": self.long_code,
            "category_1": self.category_1,
            "category_2": self.category_2,
            "unit": self.unit,
            "currency": self.currency,
            "measurement_method": self.measurement_method,
        }


def _clean(value: Any) -> Any:
    if value == "-":
        return None
    return value


def _as_int(value: Any) -> int | None:
    if value is None:
        return None
    return int(value)


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def _dict(row: sqlite3.Row) -> dict[str, Any]:
    return dict(row)


def _alias_key(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value.strip())
    ascii_value = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    return " ".join(ascii_value.upper().split())


def _pct_change(start: float | None, end: float | None) -> float | None:
    if start in (None, 0) or end is None:
        return None
    return (float(end) - float(start)) / abs(float(start))


def _cagr(start: float | None, end: float | None, periods: int | None) -> float | None:
    if start is None or end is None or periods is None or periods <= 0:
        return None
    if start <= 0 or end <= 0:
        return None
    return (float(end) / float(start)) ** (1 / periods) - 1


def _summary_stats(values: list[float]) -> dict[str, Any]:
    if not values:
        return {
            "n": 0,
            "mean": None,
            "median": None,
            "min": None,
            "max": None,
        }
    return {
        "n": len(values),
        "mean": sum(values) / len(values),
        "median": statistics.median(values),
        "min": min(values),
        "max": max(values),
    }


class GHEDStore:
    """SQLite-backed query layer over the public GHED XLSX workbook."""

    def __init__(self, workbook_path: str | Path, sqlite_path: str | Path | None = None):
        self.path = Path(workbook_path)
        self.sqlite_path = Path(sqlite_path) if sqlite_path else self.path.with_suffix(".sqlite")
        self._conn: sqlite3.Connection | None = None
        self.ensure_sqlite()

    def _workbook(self):
        return load_workbook(self.path, read_only=True, data_only=True)

    def _connect(self) -> sqlite3.Connection:
        if self._conn is None:
            self._conn = sqlite3.connect(self.sqlite_path)
            self._conn.row_factory = sqlite3.Row
        return self._conn

    def close(self) -> None:
        if self._conn is not None:
            self._conn.close()
            self._conn = None

    def _source_signature(self) -> dict[str, Any]:
        stat = self.path.stat()
        return {
            "schema_version": SCHEMA_VERSION,
            "workbook_path": str(self.path),
            "workbook_size_bytes": stat.st_size,
            "workbook_mtime_ns": stat.st_mtime_ns,
        }

    def _stored_manifest(self) -> dict[str, Any] | None:
        if not self.sqlite_path.exists():
            return None
        try:
            conn = sqlite3.connect(self.sqlite_path)
            row = conn.execute(
                "select value from manifest where key = 'source_signature'"
            ).fetchone()
            conn.close()
        except sqlite3.Error:
            return None
        if not row:
            return None
        try:
            return json.loads(row[0])
        except json.JSONDecodeError:
            return None

    def ensure_sqlite(self) -> None:
        """Build or reuse the derived SQLite cache."""
        signature = self._source_signature()
        if self._stored_manifest() == signature:
            return
        self.rebuild_sqlite(signature)

    def rebuild_sqlite(self, signature: dict[str, Any] | None = None) -> None:
        """Rebuild the derived SQLite database from the cached workbook."""
        self.close()
        signature = signature or self._source_signature()
        self.sqlite_path.parent.mkdir(parents=True, exist_ok=True)
        tmp = self.sqlite_path.with_suffix(".sqlite.tmp")
        if tmp.exists():
            tmp.unlink()

        logger.info("Rebuilding SQLite cache from %s", self.path)
        started = time.monotonic()
        conn = sqlite3.connect(tmp)
        try:
            conn.execute("pragma journal_mode = off")
            conn.execute("pragma synchronous = off")
            conn.execute("pragma temp_store = memory")
            self._create_schema(conn)
            self._load_workbook_into(conn, signature)
            conn.commit()
        except Exception:
            conn.close()
            if tmp.exists():
                tmp.unlink()
            raise
        counts = {
            "countries": conn.execute("select count(*) from countries").fetchone()[0],
            "indicators": conn.execute("select count(*) from indicators").fetchone()[0],
            "observations": conn.execute("select count(*) from observations").fetchone()[0],
        }
        conn.close()
        tmp.replace(self.sqlite_path)
        elapsed = time.monotonic() - started
        logger.info(
            "Rebuilt SQLite cache in %.1fs (countries=%d, indicators=%d, observations=%d)",
            elapsed,
            counts["countries"],
            counts["indicators"],
            counts["observations"],
        )

    def _create_schema(self, conn: sqlite3.Connection) -> None:
        conn.executescript(
            """
            create table manifest (
                key text primary key,
                value text not null
            );
            create table countries (
                country_code text primary key,
                country_name text not null,
                region text,
                income text
            );
            create table indicators (
                indicator_code text primary key,
                indicator_name text,
                long_code text,
                category_1 text,
                category_2 text,
                unit text,
                currency text,
                measurement_method text
            );
            create table observations (
                country_code text not null,
                year integer not null,
                indicator_code text not null,
                value real not null,
                primary key (indicator_code, country_code, year)
            );
            create table metadata (
                country_code text not null,
                indicator_code text not null,
                country_name text,
                region text,
                income text,
                long_code text,
                indicator_name text,
                sources text,
                comments text,
                data_type text,
                methods_of_estimation text,
                country_footnote text,
                primary key (country_code, indicator_code)
            );
            create table version_lines (
                line_no integer primary key,
                text text not null
            );
            create index idx_observations_country_year
                on observations(country_code, year);
            create index idx_metadata_indicator
                on metadata(indicator_code);
            """
        )

    def _load_workbook_into(self, conn: sqlite3.Connection, signature: dict[str, Any]) -> None:
        wb = self._workbook()
        try:
            header = [
                str(v) for v in next(wb["Data"].iter_rows(values_only=True))
            ]
            indicators = self._load_indicators(conn, wb)
            self._load_data(conn, wb, header, set(indicators))
            self._load_metadata(conn, wb)
            self._load_version(conn, wb)
        finally:
            wb.close()

        conn.execute(
            "insert into manifest(key, value) values (?, ?)",
            ("source_signature", json.dumps(signature, sort_keys=True)),
        )
        conn.execute(
            "insert into manifest(key, value) values (?, ?)",
            ("built_at", _utc_now()),
        )

    def _load_indicators(self, conn: sqlite3.Connection, wb) -> list[str]:
        out = []
        rows = wb["Codebook"].iter_rows(values_only=True)
        next(rows)
        batch = []
        for row in rows:
            code = row[0]
            if not code or code in CORE_COLUMNS:
                continue
            out.append(str(code))
            batch.append((
                str(code),
                _clean(row[1]),
                _clean(row[2]),
                _clean(row[3]),
                _clean(row[4]),
                _clean(row[5]),
                _clean(row[6]),
                _clean(row[7]),
            ))
        conn.executemany(
            """
            insert into indicators(
                indicator_code, indicator_name, long_code, category_1,
                category_2, unit, currency, measurement_method
            ) values (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            batch,
        )
        return out

    def _load_data(
        self,
        conn: sqlite3.Connection,
        wb,
        header: list[str],
        indicator_codes: set[str],
    ) -> None:
        indicator_indices = [
            (idx, code) for idx, code in enumerate(header) if code in indicator_codes
        ]
        countries_seen: set[str] = set()
        country_batch = []
        observation_batch = []
        rows = wb["Data"].iter_rows(values_only=True)
        next(rows)
        for row in rows:
            country_code = row[1]
            if not country_code:
                continue
            country_code = str(country_code)
            if country_code not in countries_seen:
                countries_seen.add(country_code)
                country_batch.append((country_code, row[0], row[2], row[3]))
            year = _as_int(row[4])
            if year is None:
                continue
            for idx, indicator_code in indicator_indices:
                if idx >= len(row):
                    continue
                value = row[idx]
                if value is None:
                    continue
                observation_batch.append((country_code, year, indicator_code, float(value)))
                if len(observation_batch) >= 50000:
                    self._insert_observations(conn, observation_batch)
                    observation_batch.clear()
        conn.executemany(
            """
            insert into countries(country_code, country_name, region, income)
            values (?, ?, ?, ?)
            """,
            country_batch,
        )
        if observation_batch:
            self._insert_observations(conn, observation_batch)

    def _insert_observations(
        self,
        conn: sqlite3.Connection,
        rows: list[tuple[str, int, str, float]],
    ) -> None:
        conn.executemany(
            """
            insert into observations(country_code, year, indicator_code, value)
            values (?, ?, ?, ?)
            """,
            rows,
        )

    def _load_metadata(self, conn: sqlite3.Connection, wb) -> None:
        rows = wb["Metadata"].iter_rows(values_only=True)
        next(rows)
        batch = []
        for row in rows:
            if not row[1] or not row[4]:
                continue
            batch.append((
                row[1],
                row[4],
                row[0],
                row[2],
                row[3],
                row[5],
                row[6],
                row[7],
                row[8],
                row[9],
                row[10],
                row[11],
            ))
        conn.executemany(
            """
            insert or replace into metadata(
                country_code, indicator_code, country_name, region, income,
                long_code, indicator_name, sources, comments, data_type,
                methods_of_estimation, country_footnote
            ) values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            batch,
        )

    def _load_version(self, conn: sqlite3.Connection, wb) -> None:
        batch = [
            (idx, str(row[0]))
            for idx, row in enumerate(wb["Version"].iter_rows(values_only=True), start=1)
            if row and row[0]
        ]
        conn.executemany(
            "insert into version_lines(line_no, text) values (?, ?)",
            batch,
        )

    def cache_status(self) -> dict[str, Any]:
        signature = self._source_signature()
        manifest = self._stored_manifest()
        conn = self._connect()
        counts = {
            "countries": conn.execute("select count(*) from countries").fetchone()[0],
            "indicators": conn.execute("select count(*) from indicators").fetchone()[0],
            "observations": conn.execute("select count(*) from observations").fetchone()[0],
            "metadata_rows": conn.execute("select count(*) from metadata").fetchone()[0],
        }
        built = conn.execute("select value from manifest where key = 'built_at'").fetchone()
        return {
            "workbook_path": str(self.path),
            "workbook_size_bytes": signature["workbook_size_bytes"],
            "workbook_mtime_ns": signature["workbook_mtime_ns"],
            "sqlite_path": str(self.sqlite_path),
            "sqlite_exists": self.sqlite_path.exists(),
            "sqlite_size_bytes": (
                self.sqlite_path.stat().st_size if self.sqlite_path.exists() else None
            ),
            "sqlite_built_at": built[0] if built else None,
            "sqlite_current": manifest == signature,
            "schema_version": SCHEMA_VERSION,
            "counts": counts,
        }

    def indicators(
        self,
        *,
        category_1: str | None = None,
        category_2: str | None = None,
        skip: int = 0,
        top: int | None = None,
    ) -> list[dict[str, Any]]:
        conn = self._connect()
        where = []
        params: list[Any] = []
        if category_1:
            where.append("category_1 = ?")
            params.append(category_1)
        if category_2:
            where.append("category_2 = ?")
            params.append(category_2)
        sql = """
            select indicator_code, indicator_name, long_code, category_1,
                   category_2, unit, currency, measurement_method
            from indicators
        """
        if where:
            sql += " where " + " and ".join(where)
        sql += " order by indicator_code"
        if top is not None:
            sql += " limit ? offset ?"
            params.extend([top, skip])
        rows = conn.execute(sql, params).fetchall()
        return [_dict(row) for row in rows]

    def indicator_map(self) -> dict[str, Indicator]:
        return {
            row["indicator_code"]: Indicator(**row)
            for row in self.indicators()
        }

    def get_indicator(self, indicator_code: str) -> dict[str, Any] | None:
        conn = self._connect()
        row = conn.execute(
            """
            select indicator_code, indicator_name, long_code, category_1,
                   category_2, unit, currency, measurement_method
            from indicators
            where indicator_code = ?
            """,
            (indicator_code,),
        ).fetchone()
        return _dict(row) if row else None

    def indicator_count(
        self,
        *,
        category_1: str | None = None,
        category_2: str | None = None,
    ) -> int:
        where = []
        params: list[Any] = []
        if category_1:
            where.append("category_1 = ?")
            params.append(category_1)
        if category_2:
            where.append("category_2 = ?")
            params.append(category_2)
        sql = "select count(*) from indicators"
        if where:
            sql += " where " + " and ".join(where)
        return self._connect().execute(sql, params).fetchone()[0]

    def indicator_categories(self) -> dict[str, Any]:
        conn = self._connect()
        by_category_1 = [
            {"category_1": row["category_1"], "count": row["count"]}
            for row in conn.execute(
                """
                select category_1, count(*) as count
                from indicators
                group by category_1
                order by count desc, category_1
                """
            ).fetchall()
        ]
        by_category_2 = [
            {
                "category_1": row["category_1"],
                "category_2": row["category_2"],
                "count": row["count"],
            }
            for row in conn.execute(
                """
                select category_1, category_2, count(*) as count
                from indicators
                group by category_1, category_2
                order by count desc, category_1, category_2
                """
            ).fetchall()
        ]
        return {"category_1": by_category_1, "category_2": by_category_2}

    def search_indicators(
        self,
        query: str,
        top: int = 50,
        *,
        category_1: str | None = None,
        category_2: str | None = None,
    ) -> list[dict[str, Any]]:
        needle = query.lower().strip()
        if not needle:
            return []
        like = f"%{needle}%"
        where = [
            """(
                lower(coalesce(indicator_code, '')) like ?
                or lower(coalesce(indicator_name, '')) like ?
                or lower(coalesce(long_code, '')) like ?
                or lower(coalesce(category_1, '')) like ?
                or lower(coalesce(category_2, '')) like ?
                or lower(coalesce(unit, '')) like ?
                or lower(coalesce(currency, '')) like ?
            )"""
        ]
        params: list[Any] = [like, like, like, like, like, like, like]
        if category_1:
            where.append("category_1 = ?")
            params.append(category_1)
        if category_2:
            where.append("category_2 = ?")
            params.append(category_2)
        params.append(top)
        conn = self._connect()
        rows = conn.execute(
            f"""
            select indicator_code, indicator_name, long_code, category_1,
                   category_2, unit, currency, measurement_method
            from indicators
            where {" and ".join(where)}
            order by indicator_code
            limit ?
            """,
            params,
        ).fetchall()
        return [_dict(row) for row in rows]

    def available_regions(self) -> list[str]:
        rows = self._connect().execute(
            "select distinct region from countries where region is not null order by region"
        ).fetchall()
        return [row["region"] for row in rows]

    def available_incomes(self) -> list[str]:
        rows = self._connect().execute(
            "select distinct income from countries where income is not null order by income"
        ).fetchall()
        return [row["income"] for row in rows]

    def normalize_region(self, value: str | None) -> str | None:
        if value is None:
            return None
        if not str(value).strip():
            return None
        available = self.available_regions()
        canonical = REGION_ALIASES.get(_norm_group_key(value))
        if canonical and canonical in available:
            return canonical
        for region in available:
            if _norm_group_key(region) == _norm_group_key(value):
                return region
        raise ValueError(
            f"Unknown region '{value}'. Available regions: "
            f"{', '.join(available) if available else '(none in workbook)'}."
        )

    def normalize_income(self, value: str | None) -> list[str] | None:
        """Resolve an income filter to one or more canonical GHED labels.

        Single-group aliases like "Upper-middle" or "UMIC" return a one-item
        list. Collective aliases like "LMIC" or "MIC" expand to multiple
        labels (Low + Lower-middle + Upper-middle, etc.) so a single SQL
        IN-filter covers the whole group.
        """
        if value is None:
            return None
        if not str(value).strip():
            return None
        available = self.available_incomes()
        key = _norm_group_key(value)

        collective = INCOME_GROUP_ALIASES.get(key)
        if collective:
            resolved = [item for item in collective if item in available]
            if not resolved:
                raise ValueError(
                    f"Income alias '{value}' expands to {collective} but none "
                    f"of these match the workbook. Available: "
                    f"{', '.join(available) if available else '(none)'}."
                )
            return resolved

        canonical = INCOME_ALIASES.get(key)
        if canonical and canonical in available:
            return [canonical]
        for income in available:
            if _norm_group_key(income) == key:
                return [income]
        collective_hint = (
            "LMIC (Low + Lower-middle + Upper-middle), "
            "MIC (Lower-middle + Upper-middle)"
        )
        raise ValueError(
            f"Unknown income group '{value}'. Available income groups: "
            f"{', '.join(available) if available else '(none in workbook)'}. "
            f"Collective aliases: {collective_hint}."
        )

    @staticmethod
    def _income_clause(incomes: list[str]) -> tuple[str, list[str]]:
        """Build a SQL fragment matching one or many income labels."""
        if len(incomes) == 1:
            return "= ?", list(incomes)
        placeholders = ", ".join("?" for _ in incomes)
        return f"IN ({placeholders})", list(incomes)

    def countries(
        self,
        *,
        region: str | None = None,
        income: str | None = None,
    ) -> list[dict[str, Any]]:
        region = self.normalize_region(region)
        incomes = self.normalize_income(income)
        conn = self._connect()
        where = []
        params: list[Any] = []
        if region:
            where.append("region = ?")
            params.append(region)
        if incomes:
            clause, values = self._income_clause(incomes)
            where.append(f"income {clause}")
            params.extend(values)
        sql = """
            select country_code, country_name, region, income
            from countries
        """
        if where:
            sql += " where " + " and ".join(where)
        sql += " order by country_name"
        rows = conn.execute(
            sql,
            params,
        ).fetchall()
        return [_dict(row) for row in rows]

    def country_groups(self) -> dict[str, Any]:
        conn = self._connect()
        regions = [
            {"region": row["region"], "country_count": row["country_count"]}
            for row in conn.execute(
                """
                select region, count(*) as country_count
                from countries
                where region is not null
                group by region
                order by region
                """
            ).fetchall()
        ]
        incomes = [
            {"income": row["income"], "country_count": row["country_count"]}
            for row in conn.execute(
                """
                select income, count(*) as country_count
                from countries
                where income is not null
                group by income
                order by income
                """
            ).fetchall()
        ]
        cross = [
            {
                "region": row["region"],
                "income": row["income"],
                "country_count": row["country_count"],
            }
            for row in conn.execute(
                """
                select region, income, count(*) as country_count
                from countries
                where region is not null and income is not null
                group by region, income
                order by region, income
                """
            ).fetchall()
        ]
        return {"regions": regions, "income_groups": incomes, "region_income": cross}

    def country_map(self) -> dict[str, dict[str, Any]]:
        return {row["country_code"]: row for row in self.countries()}

    def find_countries(self, query: str) -> list[dict[str, Any]]:
        needle = query.lower().strip()
        like = f"%{needle}%"
        conn = self._connect()
        alias_code = COUNTRY_ALIASES.get(_alias_key(query))
        alias_clause = " or country_code = ?" if alias_code else ""
        params: list[Any] = [like, like]
        if alias_code:
            params.append(alias_code)
        rows = conn.execute(
            f"""
            select country_code, country_name, region, income
            from countries
            where lower(country_code) like ? or lower(country_name) like ?{alias_clause}
            order by country_name
            """,
            params,
        ).fetchall()
        return [_dict(row) for row in rows]

    def resolve_country(self, country: str) -> str:
        value = country.strip()
        if not value:
            raise ValueError("Country cannot be empty.")
        upper = value.upper()
        conn = self._connect()
        alias_code = COUNTRY_ALIASES.get(_alias_key(value))
        if alias_code:
            row = conn.execute(
                "select country_code from countries where country_code = ?",
                (alias_code,),
            ).fetchone()
            if row:
                return row["country_code"]

        row = conn.execute(
            "select country_code from countries where country_code = ?",
            (upper,),
        ).fetchone()
        if row:
            return row["country_code"]

        matches = conn.execute(
            """
            select country_code, country_name
            from countries
            where lower(country_name) like ?
            order by country_name
            """,
            (f"%{value.lower()}%",),
        ).fetchall()
        if len(matches) == 1:
            return matches[0]["country_code"]
        if len(matches) > 1:
            names = ", ".join(row["country_name"] for row in matches[:5])
            more = "..." if len(matches) > 5 else ""
            raise ValueError(
                f"Ambiguous country '{country}': {len(matches)} matches "
                f"({names}{more}). Pass an ISO3 code."
            )
        raise ValueError(f"Could not resolve country '{country}'. Pass an ISO3 code.")

    def version(self) -> dict[str, Any]:
        conn = self._connect()
        rows = conn.execute(
            "select text from version_lines order by line_no"
        ).fetchall()
        return {"lines": [row["text"] for row in rows]}

    def indicator_data(
        self,
        indicator_code: str,
        *,
        countries: Iterable[str] | None = None,
        region: str | None = None,
        income: str | None = None,
        year_start: int | None = None,
        year_end: int | None = None,
        latest_only: bool = False,
        top: int = 1000,
    ) -> list[dict[str, Any]]:
        if self.get_indicator(indicator_code) is None:
            raise ValueError(f"Unknown GHED indicator '{indicator_code}'.")
        region = self.normalize_region(region)
        incomes = self.normalize_income(income)

        resolved = None
        if countries:
            resolved = [self.resolve_country(c) for c in countries]

        where = ["o.indicator_code = ?"]
        params: list[Any] = [indicator_code]
        if resolved:
            placeholders = ", ".join("?" for _ in resolved)
            where.append(f"o.country_code in ({placeholders})")
            params.extend(resolved)
        if region:
            where.append("c.region = ?")
            params.append(region)
        if incomes:
            clause, values = self._income_clause(incomes)
            where.append(f"c.income {clause}")
            params.extend(values)
        if year_start is not None:
            where.append("o.year >= ?")
            params.append(int(year_start))
        if year_end is not None:
            where.append("o.year <= ?")
            params.append(int(year_end))

        if latest_only:
            query = f"""
                with ranked as (
                    select
                        o.indicator_code,
                        i.indicator_name,
                        o.country_code,
                        c.country_name,
                        c.region,
                        c.income,
                        o.year,
                        o.value,
                        i.unit,
                        i.currency,
                        i.category_1,
                        i.category_2,
                        row_number() over (
                            partition by o.country_code
                            order by o.year desc
                        ) as rn
                    from observations o
                    join countries c on c.country_code = o.country_code
                    left join indicators i on i.indicator_code = o.indicator_code
                    where {" and ".join(where)}
                )
                select indicator_code, indicator_name, country_code, country_name,
                       region, income, year, value, unit, currency, category_1,
                       category_2
                from ranked
                where rn = 1
                order by country_code
                limit ?
            """
        else:
            query = f"""
                select
                    o.indicator_code,
                    i.indicator_name,
                    o.country_code,
                    c.country_name,
                    c.region,
                    c.income,
                    o.year,
                    o.value,
                    i.unit,
                    i.currency,
                    i.category_1,
                    i.category_2
                from observations o
                join countries c on c.country_code = o.country_code
                left join indicators i on i.indicator_code = o.indicator_code
                where {" and ".join(where)}
                order by o.country_code, o.year desc
                limit ?
            """
        params.append(top)
        rows = self._connect().execute(query, params).fetchall()
        return [_dict(row) for row in rows]

    def data_availability(
        self,
        indicator_codes: list[str],
        *,
        countries: Iterable[str] | None = None,
        region: str | None = None,
        income: str | None = None,
        year_start: int | None = None,
        year_end: int | None = None,
    ) -> list[dict[str, Any]]:
        for code in indicator_codes:
            if self.get_indicator(code) is None:
                raise ValueError(f"Unknown GHED indicator '{code}'.")
        region = self.normalize_region(region)
        incomes = self.normalize_income(income)
        resolved = None
        if countries:
            resolved = [self.resolve_country(c) for c in countries]

        where = [f"o.indicator_code in ({', '.join('?' for _ in indicator_codes)})"]
        params: list[Any] = list(indicator_codes)
        if resolved:
            where.append(f"o.country_code in ({', '.join('?' for _ in resolved)})")
            params.extend(resolved)
        if region:
            where.append("c.region = ?")
            params.append(region)
        if incomes:
            clause, values = self._income_clause(incomes)
            where.append(f"c.income {clause}")
            params.extend(values)
        if year_start is not None:
            where.append("o.year >= ?")
            params.append(int(year_start))
        if year_end is not None:
            where.append("o.year <= ?")
            params.append(int(year_end))

        rows = self._connect().execute(
            f"""
            select
                o.indicator_code,
                i.indicator_name,
                count(*) as observation_count,
                count(distinct o.country_code) as country_count,
                min(o.year) as first_year,
                max(o.year) as latest_year
            from observations o
            join countries c on c.country_code = o.country_code
            left join indicators i on i.indicator_code = o.indicator_code
            where {" and ".join(where)}
            group by o.indicator_code, i.indicator_name
            order by o.indicator_code
            """,
            params,
        ).fetchall()
        found = {row["indicator_code"] for row in rows}
        out = [_dict(row) for row in rows]
        for code in indicator_codes:
            if code not in found:
                indicator = self.get_indicator(code) or {}
                out.append({
                    "indicator_code": code,
                    "indicator_name": indicator.get("indicator_name"),
                    "observation_count": 0,
                    "country_count": 0,
                    "first_year": None,
                    "latest_year": None,
                })
        return out

    def research_panel(
        self,
        indicator_codes: list[str],
        *,
        countries: Iterable[str] | None = None,
        region: str | None = None,
        income: str | None = None,
        year_start: int | None = None,
        year_end: int | None = None,
        top: int = 10000,
    ) -> list[dict[str, Any]]:
        for code in indicator_codes:
            if self.get_indicator(code) is None:
                raise ValueError(f"Unknown GHED indicator '{code}'.")
        region = self.normalize_region(region)
        incomes = self.normalize_income(income)
        resolved = None
        if countries:
            resolved = [self.resolve_country(c) for c in countries]

        where = [f"o.indicator_code in ({', '.join('?' for _ in indicator_codes)})"]
        params: list[Any] = list(indicator_codes)
        if resolved:
            where.append(f"o.country_code in ({', '.join('?' for _ in resolved)})")
            params.extend(resolved)
        if region:
            where.append("c.region = ?")
            params.append(region)
        if incomes:
            clause, values = self._income_clause(incomes)
            where.append(f"c.income {clause}")
            params.extend(values)
        if year_start is not None:
            where.append("o.year >= ?")
            params.append(int(year_start))
        if year_end is not None:
            where.append("o.year <= ?")
            params.append(int(year_end))
        params.append(top)

        rows = self._connect().execute(
            f"""
            select
                o.indicator_code,
                i.indicator_name,
                o.country_code,
                c.country_name,
                c.region,
                c.income,
                o.year,
                o.value,
                i.unit,
                i.currency,
                i.category_1,
                i.category_2
            from observations o
            join countries c on c.country_code = o.country_code
            left join indicators i on i.indicator_code = o.indicator_code
            where {" and ".join(where)}
            order by o.country_code, o.year, o.indicator_code
            limit ?
            """,
            params,
        ).fetchall()
        return [_dict(row) for row in rows]

    def country_profile(
        self,
        country: str,
        *,
        year: int | None = None,
        indicator_codes: list[str] | None = None,
    ) -> list[dict[str, Any]]:
        code = self.resolve_country(country)
        codes = indicator_codes or DEFAULT_PROFILE_INDICATORS
        missing = [
            indicator_code
            for indicator_code in codes
            if self.get_indicator(indicator_code) is None
        ]
        if missing:
            raise ValueError(f"Unknown GHED indicator(s): {', '.join(missing)}")

        placeholders = ", ".join("?" for _ in codes)
        where = [
            "o.country_code = ?",
            f"o.indicator_code in ({placeholders})",
        ]
        params: list[Any] = [code, *codes]
        if year is not None:
            where.append("o.year <= ?")
            params.append(int(year))

        query = f"""
            with ranked as (
                select
                    o.indicator_code,
                    i.indicator_name,
                    o.country_code,
                    c.country_name,
                    c.region,
                    c.income,
                    o.year,
                    o.value,
                    i.unit,
                    i.currency,
                    i.category_1,
                    i.category_2,
                    row_number() over (
                        partition by o.indicator_code
                        order by o.year desc
                    ) as rn
                from observations o
                join countries c on c.country_code = o.country_code
                left join indicators i on i.indicator_code = o.indicator_code
                where {" and ".join(where)}
            )
            select indicator_code, indicator_name, country_code, country_name,
                   region, income, year, value, unit, currency, category_1,
                   category_2
            from ranked
            where rn = 1
        """
        rows = {
            row["indicator_code"]: _dict(row)
            for row in self._connect().execute(query, params).fetchall()
        }
        countries = self.country_map()
        results = []
        for indicator_code in codes:
            if indicator_code in rows:
                results.append(rows[indicator_code])
            else:
                indicator = self.get_indicator(indicator_code) or {}
                results.append({
                    "indicator_code": indicator_code,
                    "indicator_name": indicator.get("indicator_name"),
                    "country_code": code,
                    "country_name": countries[code]["country_name"],
                    "year": None,
                    "value": None,
                })
        return results

    def country_metadata(
        self,
        *,
        country: str | None = None,
        indicator_code: str | None = None,
        top: int = 20,
    ) -> list[dict[str, Any]]:
        where = []
        params: list[Any] = []
        if country:
            where.append("country_code = ?")
            params.append(self.resolve_country(country))
        if indicator_code:
            where.append("indicator_code = ?")
            params.append(indicator_code)
        sql = """
            select country_name, country_code, region, income, indicator_code,
                   long_code, indicator_name, sources, comments, data_type,
                   methods_of_estimation, country_footnote
            from metadata
        """
        if where:
            sql += " where " + " and ".join(where)
        sql += " order by country_code, indicator_code limit ?"
        params.append(top)
        rows = self._connect().execute(sql, params).fetchall()
        return [_dict(row) for row in rows]

    def group_summary(
        self,
        indicator_code: str,
        *,
        region: str | None = None,
        income: str | None = None,
        year: int | None = None,
        latest_only: bool = True,
        top_n: int = 5,
    ) -> dict[str, Any]:
        if not region and not income:
            raise ValueError("Pass at least one of 'region' or 'income'.")
        if year is not None:
            rows = self.indicator_data(
                indicator_code,
                region=region,
                income=income,
                year_start=year,
                year_end=year,
                latest_only=False,
                top=10000,
            )
        else:
            rows = self.indicator_data(
                indicator_code,
                region=region,
                income=income,
                latest_only=latest_only,
                top=10000,
            )
        values = [float(row["value"]) for row in rows if row.get("value") is not None]
        sorted_rows = sorted(
            rows,
            key=lambda row: (
                row.get("value") is None,
                row.get("value") if row.get("value") is not None else 0,
            ),
        )
        years = sorted({row["year"] for row in rows if row.get("year") is not None})
        countries_in_group = self.countries(region=region, income=income)
        covered = {row["country_code"] for row in rows}
        return {
            "indicator": self.get_indicator(indicator_code),
            "region": region,
            "income": income,
            "year": year,
            "latest_only": latest_only if year is None else False,
            "country_count": len(countries_in_group),
            "countries_with_data": len(covered),
            "coverage_ratio": (
                None if not countries_in_group else len(covered) / len(countries_in_group)
            ),
            "stats": _summary_stats(values),
            "years_seen": years,
            "bottom": sorted_rows[:top_n],
            "top": list(reversed(sorted_rows[-top_n:])),
            "warnings": (
                [{
                    "type": "mixed_latest_years",
                    "message": (
                        "The group summary uses different latest years across "
                        "countries. Interpret ranks and summary statistics with care."
                    ),
                    "years_seen": years,
                }] if latest_only and year is None and len(years) > 1 else []
            ),
        }

    def indicator_trends(
        self,
        indicator_code: str,
        *,
        countries: Iterable[str] | None = None,
        region: str | None = None,
        income: str | None = None,
        year_start: int | None = None,
        year_end: int | None = None,
        min_year_count: int | None = None,
        min_period_years: int | None = None,
        top: int = 1000,
    ) -> list[dict[str, Any]]:
        if self.get_indicator(indicator_code) is None:
            raise ValueError(f"Unknown GHED indicator '{indicator_code}'.")
        region = self.normalize_region(region)
        incomes = self.normalize_income(income)
        resolved = None
        if countries:
            resolved = [self.resolve_country(c) for c in countries]

        where = ["o.indicator_code = ?"]
        params: list[Any] = [indicator_code]
        if resolved:
            where.append(f"o.country_code in ({', '.join('?' for _ in resolved)})")
            params.extend(resolved)
        if region:
            where.append("c.region = ?")
            params.append(region)
        if incomes:
            clause, values = self._income_clause(incomes)
            where.append(f"c.income {clause}")
            params.extend(values)
        if year_start is not None:
            where.append("o.year >= ?")
            params.append(int(year_start))
        if year_end is not None:
            where.append("o.year <= ?")
            params.append(int(year_end))

        rows = self._connect().execute(
            f"""
            select o.country_code, c.country_name, c.region, c.income,
                   o.year, o.value
            from observations o
            join countries c on c.country_code = o.country_code
            where {" and ".join(where)}
            order by o.country_code, o.year
            """,
            params,
        ).fetchall()
        by_country: dict[str, list[dict[str, Any]]] = {}
        for row in rows:
            payload = _dict(row)
            by_country.setdefault(payload["country_code"], []).append(payload)

        out = []
        for code, series in by_country.items():
            first = series[0]
            last = series[-1]
            period_years = int(last["year"]) - int(first["year"])
            if min_year_count is not None and len(series) < int(min_year_count):
                continue
            if min_period_years is not None and period_years < int(min_period_years):
                continue
            out.append({
                "country_code": code,
                "country_name": first["country_name"],
                "region": first["region"],
                "income": first["income"],
                "first_year": first["year"],
                "first_value": first["value"],
                "latest_year": last["year"],
                "latest_value": last["value"],
                "year_count": len(series),
                "period_years": period_years,
                "absolute_change": float(last["value"]) - float(first["value"]),
                "percent_change": _pct_change(first["value"], last["value"]),
                "cagr": _cagr(first["value"], last["value"], period_years),
            })
        out.sort(key=lambda row: row["absolute_change"], reverse=True)
        return out[:top]

    def quality_assessment(
        self,
        indicator_code: str,
        *,
        country: str | None = None,
        countries: Iterable[str] | None = None,
        region: str | None = None,
        income: str | None = None,
        year_start: int | None = None,
        year_end: int | None = None,
        top: int = 100,
    ) -> dict[str, Any]:
        if self.get_indicator(indicator_code) is None:
            raise ValueError(f"Unknown GHED indicator '{indicator_code}'.")
        requested = list(countries or [])
        if country:
            requested.append(country)
        group_countries = self.countries(region=region, income=income)
        if requested:
            target_codes = [self.resolve_country(item) for item in requested]
        elif region or income:
            target_codes = [row["country_code"] for row in group_countries]
        else:
            target_codes = []

        metadata_rows = self.country_metadata(
            indicator_code=indicator_code,
            top=100000,
        )
        if target_codes:
            target_set = set(target_codes)
            metadata_rows = [
                row for row in metadata_rows if row["country_code"] in target_set
            ]
        data_types: dict[str, int] = {}
        source_count = 0
        estimation_notes = 0
        footnotes = 0
        comments = 0
        for row in metadata_rows:
            data_type = row.get("data_type") or "Unspecified"
            data_types[data_type] = data_types.get(data_type, 0) + 1
            source_count += 1 if row.get("sources") else 0
            estimation_notes += 1 if row.get("methods_of_estimation") else 0
            footnotes += 1 if row.get("country_footnote") else 0
            comments += 1 if row.get("comments") else 0

        availability = self.data_availability(
            [indicator_code],
            countries=target_codes or None,
            region=region if not target_codes else None,
            income=income if not target_codes else None,
            year_start=year_start,
            year_end=year_end,
        )[0]
        warnings = []
        if availability["observation_count"] == 0:
            warnings.append({
                "type": "no_observations",
                "message": "No observations matched the requested filters.",
            })
        if data_types and any("estimated" in k.lower() for k in data_types):
            warnings.append({
                "type": "estimated_data_present",
                "message": "At least one metadata row is marked as estimated.",
            })
        if footnotes or comments:
            warnings.append({
                "type": "country_notes_present",
                "message": "Country footnotes or comments are present; inspect rows before strong claims.",
            })

        return {
            "indicator": self.get_indicator(indicator_code),
            "country": country,
            "countries": list(countries or []),
            "region": region,
            "income": income,
            "year_start": year_start,
            "year_end": year_end,
            "availability": availability,
            "metadata_summary": {
                "metadata_rows": len(metadata_rows),
                "data_types": data_types,
                "rows_with_sources": source_count,
                "rows_with_estimation_notes": estimation_notes,
                "rows_with_country_footnotes": footnotes,
                "rows_with_comments": comments,
            },
            "sample_metadata_rows": metadata_rows[:top],
            "warnings": warnings,
        }

    def additive_relationships(self, indicator_code: str) -> list[dict[str, Any]]:
        """Return known additive child relationships for a variable."""
        parent = self.get_indicator(indicator_code)
        if parent is None:
            raise ValueError(f"Unknown GHED indicator '{indicator_code}'.")

        relationships = []
        for rel in ADDITIVE_RELATIONSHIPS.get(indicator_code, []):
            relationships.append({
                **rel,
                "parent_code": indicator_code,
                "parent_name": parent.get("indicator_name"),
                "children_metadata": [
                    self.get_indicator(code) for code in rel["children"]
                ],
                "source": "curated_codebook_formula",
                "additivity": "sum(children) ~= parent for amount variables",
            })

        dynamic_children = self._sha_direct_children(indicator_code)
        if dynamic_children:
            relationships.append({
                "relationship_id": "sha_direct_children",
                "description": "Direct children in the SHA 2011 long-code hierarchy.",
                "parent_code": indicator_code,
                "parent_name": parent.get("indicator_name"),
                "children": [row["indicator_code"] for row in dynamic_children],
                "children_metadata": dynamic_children,
                "basis": "Direct child long codes under the parent's sha11.* long code.",
                "source": "inferred_from_sha11_long_codes",
                "additivity": "sum(children) ~= parent when all variables are current-NCU amount series",
            })
        return relationships

    def explain_relationship(self, indicator_code: str) -> dict[str, Any]:
        indicator = self.get_indicator(indicator_code)
        if indicator is None:
            raise ValueError(f"Unknown GHED indicator '{indicator_code}'.")
        unit = indicator.get("unit")
        currency = indicator.get("currency")
        long_code = indicator.get("long_code") or ""
        method = indicator.get("measurement_method") or ""
        is_amount = unit == "Millions" and currency == "NCU"
        is_share = (
            unit == "Percentage"
            or "%" in str(long_code)
            or "/" in str(method)
            or "per" in str(unit or "").lower()
        )
        children = self.additive_relationships(indicator_code)
        parent_candidates = []
        for parent_code, relationships in ADDITIVE_RELATIONSHIPS.items():
            for rel in relationships:
                if indicator_code in rel.get("children", []):
                    parent = self.get_indicator(parent_code) or {}
                    parent_candidates.append({
                        "parent_code": parent_code,
                        "parent_name": parent.get("indicator_name"),
                        "relationship_id": rel.get("relationship_id"),
                        "relationship_description": rel.get("description"),
                        "source": "curated_codebook_formula",
                    })
        if is_amount and str(long_code).startswith("sha11.") and "." in str(long_code):
            parent_long_code = str(long_code).rsplit(".", 1)[0]
            parent = self._connect().execute(
                """
                select indicator_code, indicator_name
                from indicators
                where long_code = ? and unit = 'Millions' and currency = 'NCU'
                """,
                (parent_long_code,),
            ).fetchone()
            if parent:
                parent_candidates.append({
                    "parent_code": parent["indicator_code"],
                    "parent_name": parent["indicator_name"],
                    "relationship_id": "sha_direct_children",
                    "relationship_description": "Direct parent in the SHA 2011 long-code hierarchy.",
                    "source": "inferred_from_sha11_long_codes",
                })

        if children:
            role = "additive_parent"
        elif parent_candidates:
            role = "component"
        elif is_share:
            role = "derived_ratio_or_share"
        elif is_amount:
            role = "amount_series"
        else:
            role = "context_or_conversion_series"

        cautions = []
        if is_share:
            cautions.append(
                "Do not sum this variable across components as an accounting identity."
            )
        if is_amount:
            cautions.append(
                "Amount series can be additive only when parent/child units, currency, prices, and classification dimensions match."
            )
        if children:
            cautions.append(
                "Use additive_hierarchy or build_additive_breakdown to validate any decomposition for a country-year."
            )
        if not children and not parent_candidates and not is_share:
            cautions.append(
                "No curated or inferred additive hierarchy is currently known for this variable."
            )

        return {
            "indicator": indicator,
            "role": role,
            "is_amount_current_ncu": is_amount,
            "is_ratio_share_or_per_capita": is_share,
            "additive_children": children,
            "known_parents": parent_candidates,
            "interpretation_cautions": cautions,
        }

    def _sha_direct_children(self, indicator_code: str) -> list[dict[str, Any]]:
        parent = self.get_indicator(indicator_code)
        if not parent or parent.get("unit") != "Millions" or parent.get("currency") != "NCU":
            return []
        long_code = parent.get("long_code")
        if not long_code or not str(long_code).startswith("sha11."):
            return []

        conn = self._connect()
        rows = conn.execute(
            """
            select indicator_code, indicator_name, long_code, category_1,
                   category_2, unit, currency, measurement_method
            from indicators
            where unit = 'Millions'
              and currency = 'NCU'
              and long_code like ?
            order by long_code
            """,
            (f"{long_code}.%",),
        ).fetchall()
        direct = []
        for row in rows:
            child = _dict(row)
            if self._is_direct_sha_child(str(long_code), str(child["long_code"])):
                direct.append(child)
        return direct

    def _is_direct_sha_child(self, parent_long_code: str, child_long_code: str) -> bool:
        if not child_long_code.startswith(parent_long_code + "."):
            return False
        rest = child_long_code[len(parent_long_code) + 1:]
        if "_" in rest:
            return False
        return "." not in rest

    def breakdown(
        self,
        parent_code: str,
        *,
        country: str,
        year: int,
        relationship_id: str | None = None,
    ) -> dict[str, Any]:
        """Build and validate an additive breakdown for a country-year."""
        relationships = self.additive_relationships(parent_code)
        if relationship_id:
            relationships = [
                rel for rel in relationships
                if rel["relationship_id"] == relationship_id
            ]
            if not relationships:
                raise ValueError(
                    f"No relationship '{relationship_id}' found for '{parent_code}'."
                )
        if not relationships:
            raise ValueError(f"No additive hierarchy known for '{parent_code}'.")
        rel = relationships[0]

        resolved_country = self.resolve_country(country)
        parent_rows = self.indicator_data(
            parent_code,
            countries=[resolved_country],
            year_start=year,
            year_end=year,
            top=1,
        )
        parent_value = parent_rows[0]["value"] if parent_rows else None
        children = []
        child_sum = 0.0
        for child_code in rel["children"]:
            rows = self.indicator_data(
                child_code,
                countries=[resolved_country],
                year_start=year,
                year_end=year,
                top=1,
            )
            row = rows[0] if rows else None
            value = row["value"] if row else None
            if value is not None:
                child_sum += float(value)
            children.append({
                "indicator_code": child_code,
                "indicator_name": (
                    (row or self.get_indicator(child_code) or {}).get("indicator_name")
                ),
                "value": value,
                "share_of_parent": (
                    None if value is None or not parent_value
                    else float(value) / float(parent_value)
                ),
            })

        difference = None if parent_value is None else child_sum - float(parent_value)
        relative_difference = (
            None if parent_value in (None, 0)
            else difference / float(parent_value)
        )
        return {
            "relationship": rel,
            "country_code": resolved_country,
            "country": self.country_map()[resolved_country]["country_name"],
            "year": year,
            "parent": {
                "indicator_code": parent_code,
                "indicator_name": (self.get_indicator(parent_code) or {}).get("indicator_name"),
                "value": parent_value,
            },
            "children": children,
            "child_sum": child_sum,
            "difference": difference,
            "relative_difference": relative_difference,
            "balanced": (
                False if relative_difference is None
                else abs(relative_difference) < 1e-6
            ),
            "caution": (
                "Only current-NCU amount variables are additive. Percentages, per-capita "
                "series, USD/PPP conversions, and constant-price variants should not be "
                "summed as accounting identities."
            ),
        }


def rows_to_csv(rows: list[dict[str, Any]]) -> str:
    if not rows:
        return ""
    fields = [
        "indicator_code",
        "indicator_name",
        "country_code",
        "country_name",
        "region",
        "income",
        "year",
        "value",
        "unit",
        "currency",
    ]
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=fields, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows)
    return buf.getvalue()


def dicts_to_csv(rows: list[dict[str, Any]]) -> str:
    if not rows:
        return ""
    fields: list[str] = []
    for row in rows:
        for key in row:
            if key not in fields:
                fields.append(key)
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=fields, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows)
    return buf.getvalue()
